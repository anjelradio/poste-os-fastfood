from decimal import Decimal
from io import BytesIO

from django.db.models import Count, Sum
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from apps.orders.models import OrderProducts


def _fmt_decimal_2(value):
    return f"{Decimal(value):.2f}"


def get_product_sales_for_report(from_date, to_date, product_id):
    return (
        OrderProducts.objects.filter(
            state=True,
            order__state=True,
            order__created_date__range=[from_date, to_date],
            product_id=product_id,
        )
        .exclude(order__status="CANCELLED")
        .select_related("order", "order__client", "product")
        .order_by("order__created_date", "order__order_number", "id")
    )


def build_product_sales_report_summary(*, sales, from_date, to_date, product):
    orders_count = sales.values("order_id").distinct().count()
    quantity_sold = sales.aggregate(total=Sum("quantity"))["total"] or Decimal("0.00")
    total_revenue = sales.aggregate(total=Sum("subtotal"))["total"] or Decimal("0.00")
    average_order_revenue = total_revenue / orders_count if orders_count else Decimal("0.00")
    best_day = (
        sales.values("order__created_date")
        .annotate(total=Sum("subtotal"), orders_count=Count("order_id", distinct=True))
        .order_by("-total", "-orders_count")
        .first()
    )

    recent_sales = []
    for sale in sales.order_by("-order__created_date", "-order__order_number", "-id")[:5]:
        order = sale.order
        recent_sales.append(
            {
                "orderNumber": order.order_number or order.id,
                "date": str(order.created_date),
                "client": order.client.name if order.client else "Sin cliente",
                "quantity": float(_fmt_decimal_2(sale.quantity)),
                "subtotal": float(_fmt_decimal_2(sale.subtotal)),
            }
        )

    return {
        "fromDate": str(from_date),
        "toDate": str(to_date),
        "product": {"id": product.id, "name": product.name},
        "ordersCount": orders_count,
        "quantitySold": float(_fmt_decimal_2(quantity_sold)),
        "totalRevenue": float(_fmt_decimal_2(total_revenue)),
        "averageOrderRevenue": float(_fmt_decimal_2(average_order_revenue)),
        "bestDay": (
            {
                "date": str(best_day["order__created_date"]),
                "total": float(_fmt_decimal_2(best_day["total"] or Decimal("0.00"))),
                "ordersCount": best_day["orders_count"],
            }
            if best_day
            else None
        ),
        "recentSales": recent_sales,
    }


def build_product_sales_report_pdf(
    *,
    sales,
    from_date,
    to_date,
    include_orders,
    generated_by,
):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        textColor=colors.HexColor("#111827"),
        alignment=1,
        spaceAfter=10,
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        textColor=colors.HexColor("#374151"),
        spaceAfter=4,
    )
    section_title_style = ParagraphStyle(
        "SectionTitle",
        parent=styles["Heading3"],
        fontName="Helvetica-Bold",
        fontSize=11,
        textColor=colors.HexColor("#111827"),
        spaceBefore=8,
        spaceAfter=6,
    )

    elements = []

    first_sale = sales.first()
    product_name = first_sale.product.name if first_sale else "-"
    orders_count = sales.values("order_id").distinct().count()
    quantity_sold = sales.aggregate(total=Sum("quantity"))["total"] or Decimal("0.00")
    total_revenue = sales.aggregate(total=Sum("subtotal"))["total"] or Decimal("0.00")

    report_rows = [
        ["Producto", product_name],
        ["Rango", f"{from_date} a {to_date}"],
        ["Generado por", generated_by],
        ["Aparece en órdenes", str(orders_count)],
        ["Cantidad vendida", _fmt_decimal_2(quantity_sold)],
        ["Total ventas", f"Bs. {_fmt_decimal_2(total_revenue)}"],
    ]

    elements.append(Paragraph("REPORTE DE VENTAS POR PRODUCTO", title_style))
    elements.append(Spacer(1, 0.2 * cm))

    info_table = Table(report_rows, colWidths=[4.5 * cm, 11.7 * cm])
    info_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F3F4F6")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#111827")),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D1D5DB")),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    elements.append(info_table)

    if include_orders:
        elements.append(Spacer(1, 0.35 * cm))
        elements.append(Paragraph("Detalle por orden", section_title_style))

        for sale in sales:
            order = sale.order
            client_name = order.client.name if order.client else "Sin cliente"
            elements.append(
                Paragraph(
                    f"<b>Orden #{order.order_number or order.id}</b> - {client_name} - {order.created_date}",
                    subtitle_style,
                )
            )

            detail_rows = [["Item", "Cantidad", "Subtotal"]]
            detail_rows.append(
                [
                    sale.product.name,
                    _fmt_decimal_2(sale.quantity),
                    f"Bs. {_fmt_decimal_2(sale.subtotal)}",
                ]
            )

            details_table = Table(
                detail_rows,
                colWidths=[9.0 * cm, 3.0 * cm, 4.2 * cm],
                repeatRows=1,
            )
            details_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#374151")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D1D5DB")),
                        ("ALIGN", (1, 1), (2, -1), "RIGHT"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("TOPPADDING", (0, 0), (-1, -1), 3),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ]
                )
            )
            elements.append(details_table)
            elements.append(Spacer(1, 0.25 * cm))

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def build_product_sales_report_excel(
    *,
    sales,
    from_date,
    to_date,
    include_orders,
    generated_by,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "VentasProducto"

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)

    first_sale = sales.first()
    product_name = first_sale.product.name if first_sale else "-"
    orders_count = sales.values("order_id").distinct().count()
    quantity_sold = sales.aggregate(total=Sum("quantity"))["total"] or Decimal("0.00")
    total_revenue = sales.aggregate(total=Sum("subtotal"))["total"] or Decimal("0.00")

    sheet["A1"] = "REPORTE DE VENTAS POR PRODUCTO"
    sheet["A1"].font = Font(size=14, bold=True)

    sheet["A3"] = "Producto"
    sheet["B3"] = product_name
    sheet["A4"] = "Rango"
    sheet["B4"] = f"{from_date} a {to_date}"
    sheet["A5"] = "Generado por"
    sheet["B5"] = generated_by
    sheet["A6"] = "Aparece en órdenes"
    sheet["B6"] = orders_count
    sheet["A7"] = "Cantidad vendida"
    sheet["B7"] = float(_fmt_decimal_2(quantity_sold))
    sheet["A8"] = "Total ventas"
    sheet["B8"] = float(_fmt_decimal_2(total_revenue))

    for cell_ref in ["A3", "A4", "A5", "A6", "A7", "A8"]:
        sheet[cell_ref].font = bold_font

    start_row = 10
    columns = ["Orden", "Fecha", "Cliente", "Item", "Cantidad", "Subtotal"]
    for idx, value in enumerate(columns, start=1):
        cell = sheet.cell(row=start_row, column=idx, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row = start_row + 1
    for sale in sales:
        order = sale.order
        client_name = order.client.name if order.client else "Sin cliente"
        sheet.cell(row=row, column=1, value=f"#{order.order_number or order.id}")
        sheet.cell(row=row, column=2, value=str(order.created_date))
        sheet.cell(row=row, column=3, value=client_name)
        sheet.cell(row=row, column=4, value=sale.product.name)
        sheet.cell(row=row, column=5, value=float(_fmt_decimal_2(sale.quantity)))
        sheet.cell(row=row, column=6, value=float(_fmt_decimal_2(sale.subtotal)))
        row += 1

        if include_orders:
            row += 0

    sheet.column_dimensions["A"].width = 12
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 30
    sheet.column_dimensions["D"].width = 24
    sheet.column_dimensions["E"].width = 12
    sheet.column_dimensions["F"].width = 14

    for current_row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=1, max_col=6):
        for cell in current_row:
            if cell.column in [5, 6] and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()
