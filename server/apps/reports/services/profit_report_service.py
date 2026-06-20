from decimal import Decimal
from io import BytesIO

from django.db.models import Count, Sum
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from apps.orders.models import Order


def _fmt_decimal_2(value):
    return f"{Decimal(value):.2f}"


def get_orders_for_profit_report(from_date, to_date):
    return (
        Order.objects.filter(
            state=True,
            created_date__range=[from_date, to_date],
        )
        .exclude(status=Order.Status.CANCELLED)
        .select_related("client", "user")
        .order_by("created_date", "order_number", "id")
    )


def build_profit_report_summary(*, orders, from_date, to_date):
    total_amount = orders.aggregate(total=Sum("total"))["total"] or Decimal("0.00")
    orders_count = orders.count()
    average_ticket = total_amount / orders_count if orders_count else Decimal("0.00")
    best_day = (
        orders.values("created_date")
        .annotate(total=Sum("total"), orders_count=Count("id"))
        .order_by("-total", "-orders_count")
        .first()
    )

    recent_orders = []
    for order in orders.order_by("-created_date", "-order_number", "-id")[:5]:
        recent_orders.append(
            {
                "orderNumber": order.order_number or order.id,
                "date": str(order.created_date),
                "client": order.client.name if order.client else "Sin cliente",
                "type": order.type,
                "status": order.status,
                "total": float(_fmt_decimal_2(order.total)),
            }
        )

    return {
        "fromDate": str(from_date),
        "toDate": str(to_date),
        "ordersCount": orders_count,
        "totalAmount": float(_fmt_decimal_2(total_amount)),
        "averageTicket": float(_fmt_decimal_2(average_ticket)),
        "bestDay": (
            {
                "date": str(best_day["created_date"]),
                "total": float(_fmt_decimal_2(best_day["total"] or Decimal("0.00"))),
                "ordersCount": best_day["orders_count"],
            }
            if best_day
            else None
        ),
        "recentOrders": recent_orders,
    }


def build_profit_report_pdf(*, orders, from_date, to_date, generated_by):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        textColor=colors.HexColor("#111827"),
        alignment=1,
        spaceAfter=10,
    )
    section_title_style = ParagraphStyle(
        "SectionTitle",
        parent=styles["Heading3"],
        fontName="Helvetica-Bold",
        fontSize=11,
        textColor=colors.HexColor("#111827"),
        spaceBefore=8,
        spaceAfter=6,
    )
    small_style = ParagraphStyle(
        "Small",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        textColor=colors.HexColor("#6B7280"),
    )

    elements = []

    total_amount = orders.aggregate(total=Sum("total"))["total"] or Decimal("0.00")
    report_rows = [
        ["Rango", f"{from_date} a {to_date}"],
        ["Generado por", generated_by],
        ["Cantidad de ordenes", str(orders.count())],
        ["Ganancia acumulada", f"Bs. {_fmt_decimal_2(total_amount)}"],
    ]

    elements.append(Paragraph("REPORTE DE GANANCIAS", title_style))
    elements.append(Spacer(1, 0.2 * cm))

    info_table = Table(report_rows, colWidths=[4.2 * cm, 12.0 * cm])
    info_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F3F4F6")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#111827")),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D1D5DB")),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    elements.append(info_table)
    elements.append(Spacer(1, 0.35 * cm))

    elements.append(Paragraph("Ordenes incluidas", section_title_style))

    table_data = [["Orden", "Fecha", "Cliente", "Tipo", "Estado", "Total"]]
    for order in orders:
        table_data.append(
            [
                f"#{order.order_number or order.id}",
                str(order.created_date),
                order.client.name if order.client else "Sin cliente",
                order.type,
                order.status,
                f"Bs. {_fmt_decimal_2(order.total)}",
            ]
        )

    orders_table = Table(
        table_data,
        colWidths=[1.9 * cm, 2.5 * cm, 4.5 * cm, 2.5 * cm, 2.6 * cm, 2.5 * cm],
        repeatRows=1,
    )
    style_rules = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (5, 1), (5, -1), "RIGHT"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D1D5DB")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for row_index in range(1, len(table_data)):
        if row_index % 2 == 0:
            style_rules.append(
                ("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#F9FAFB"))
            )

    orders_table.setStyle(TableStyle(style_rules))
    elements.append(orders_table)

    elements.append(Spacer(1, 0.35 * cm))
    elements.append(Paragraph("Sistema de Reportes", small_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def build_profit_report_excel(*, orders, from_date, to_date, generated_by):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ganancias"

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)

    sheet["A1"] = "REPORTE DE GANANCIAS"
    sheet["A1"].font = Font(size=14, bold=True)

    sheet["A3"] = "Rango"
    sheet["B3"] = f"{from_date} a {to_date}"
    sheet["A4"] = "Generado por"
    sheet["B4"] = generated_by
    sheet["A5"] = "Cantidad de ordenes"
    sheet["B5"] = orders.count()
    total_amount = orders.aggregate(total=Sum("total"))["total"] or Decimal("0.00")
    sheet["A6"] = "Ganancia acumulada"
    sheet["B6"] = float(_fmt_decimal_2(total_amount))

    for cell_ref in ["A3", "A4", "A5", "A6"]:
        sheet[cell_ref].font = bold_font

    start_row = 8
    columns = ["Orden", "Fecha", "Cliente", "Tipo", "Estado", "Total"]
    for idx, value in enumerate(columns, start=1):
        cell = sheet.cell(row=start_row, column=idx, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row = start_row + 1
    for order in orders:
        sheet.cell(row=row, column=1, value=f"#{order.order_number or order.id}")
        sheet.cell(row=row, column=2, value=str(order.created_date))
        sheet.cell(
            row=row,
            column=3,
            value=order.client.name if order.client else "Sin cliente",
        )
        sheet.cell(row=row, column=4, value=order.type)
        sheet.cell(row=row, column=5, value=order.status)
        sheet.cell(row=row, column=6, value=float(_fmt_decimal_2(order.total)))
        row += 1

    sheet.column_dimensions["A"].width = 12
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 34
    sheet.column_dimensions["D"].width = 16
    sheet.column_dimensions["E"].width = 16
    sheet.column_dimensions["F"].width = 16

    for current_row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=1, max_col=6):
        for cell in current_row:
            if cell.column == 6 and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()
