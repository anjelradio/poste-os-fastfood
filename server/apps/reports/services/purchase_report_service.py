from decimal import Decimal
from io import BytesIO

from django.db.models import Count, Sum
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from apps.purchases.models import Purchase


def _fmt_decimal_2(value):
    return f"{Decimal(value):.2f}"


def get_purchases_for_report(from_date, to_date, raw_material_id=None, supplier_id=None):
    queryset = Purchase.objects.filter(
        state=True,
        purchased_at__date__range=[from_date, to_date],
    )

    if raw_material_id is not None:
        queryset = queryset.filter(items__raw_material_id=raw_material_id).distinct()

    if supplier_id is not None:
        queryset = queryset.filter(supplier_id=supplier_id)

    return queryset.select_related("supplier").prefetch_related(
        "items", "items__raw_material", "items__raw_material__measure_unit"
    ).order_by("purchased_at", "id")


def build_purchases_report_summary(*, purchases, from_date, to_date):
    total_amount = purchases.aggregate(total=Sum("total"))["total"] or Decimal("0.00")
    purchases_count = purchases.count()
    average_purchase = total_amount / purchases_count if purchases_count else Decimal("0.00")
    top_supplier = (
        purchases.values("supplier__business_name")
        .annotate(total=Sum("total"), purchases_count=Count("id"))
        .order_by("-total", "-purchases_count")
        .first()
    )

    recent_purchases = []
    for purchase in purchases.order_by("-purchased_at", "-id")[:5]:
        recent_purchases.append(
            {
                "id": purchase.id,
                "date": purchase.purchased_at.strftime("%Y-%m-%d %H:%M"),
                "supplier": purchase.supplier.business_name if purchase.supplier else "Sin proveedor",
                "description": purchase.description or "-",
                "total": float(_fmt_decimal_2(purchase.total)),
            }
        )

    return {
        "fromDate": str(from_date),
        "toDate": str(to_date),
        "purchasesCount": purchases_count,
        "totalAmount": float(_fmt_decimal_2(total_amount)),
        "averagePurchase": float(_fmt_decimal_2(average_purchase)),
        "topSupplier": (
            {
                "name": top_supplier["supplier__business_name"] or "Sin proveedor",
                "total": float(_fmt_decimal_2(top_supplier["total"] or Decimal("0.00"))),
                "purchasesCount": top_supplier["purchases_count"],
            }
            if top_supplier
            else None
        ),
        "recentPurchases": recent_purchases,
    }


def build_purchases_report_pdf(
    *, purchases, from_date, to_date, include_items, generated_by, supplier_name=None
):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        textColor=colors.HexColor("#111827"),
        alignment=1,
        spaceAfter=10,
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        textColor=colors.HexColor("#374151"),
        spaceAfter=4,
    )
    section_title_style = ParagraphStyle(
        "SectionTitle",
        parent=styles["Heading3"],
        fontName="Helvetica-Bold",
        fontSize=11,
        textColor=colors.HexColor("#111827"),
        spaceBefore=8,
        spaceAfter=6,
    )
    small_style = ParagraphStyle(
        "Small",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        textColor=colors.HexColor("#6B7280"),
    )

    elements = []

    total_amount = purchases.aggregate(total=Sum("total"))["total"] or Decimal("0.00")
    report_rows = [
        ["Rango", f"{from_date} a {to_date}"],
        ["Proveedor", supplier_name or "Todos"],
        ["Generado por", generated_by],
        ["Cantidad de compras", str(purchases.count())],
        ["Total acumulado", f"Bs. {_fmt_decimal_2(total_amount)}"],
    ]

    elements.append(Paragraph("REPORTE DE COMPRAS", title_style))
    elements.append(Spacer(1, 0.2 * cm))

    info_table = Table(report_rows, colWidths=[4.2 * cm, 12.0 * cm])
    info_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F3F4F6")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#111827")),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D1D5DB")),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    elements.append(info_table)
    elements.append(Spacer(1, 0.35 * cm))

    elements.append(Paragraph("Compras encontradas", section_title_style))

    purchases_table_data = [["Compra", "Fecha", "Proveedor", "Descripcion", "Total"]]
    for purchase in purchases:
        purchases_table_data.append(
            [
                f"#{purchase.id}",
                purchase.purchased_at.strftime("%Y-%m-%d %H:%M"),
                purchase.supplier.business_name if purchase.supplier else "Sin proveedor",
                purchase.description or "-",
                f"Bs. {_fmt_decimal_2(purchase.total)}",
            ]
        )

    purchases_table = Table(
        purchases_table_data,
        colWidths=[1.5 * cm, 3.0 * cm, 4.8 * cm, 5.3 * cm, 2.0 * cm],
        repeatRows=1,
    )
    purchases_style_rules = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (4, 1), (4, -1), "RIGHT"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D1D5DB")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for row_index in range(1, len(purchases_table_data)):
        if row_index % 2 == 0:
            purchases_style_rules.append(
                ("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#F9FAFB"))
            )

    purchases_table.setStyle(TableStyle(purchases_style_rules))
    elements.append(purchases_table)

    if include_items:
        elements.append(Spacer(1, 0.4 * cm))
        elements.append(Paragraph("Detalle por compra", section_title_style))
        for purchase in purchases:
            elements.append(
                Paragraph(
                    (
                        f"<b>Compra #{purchase.id}</b> - "
                        f"{purchase.purchased_at.strftime('%Y-%m-%d %H:%M')} - "
                        f"{purchase.supplier.business_name if purchase.supplier else 'Sin proveedor'}"
                    ),
                    subtitle_style,
                )
            )

            detail_rows = [["Item", "Cantidad", "Unidad", "Unitario", "Subtotal"]]
            for item in purchase.items.all():
                subtotal = item.unit_price * item.quantity
                detail_rows.append(
                    [
                        item.raw_material.name,
                        _fmt_decimal_2(item.quantity),
                        item.raw_material.measure_unit.name,
                        f"Bs. {_fmt_decimal_2(item.unit_price)}",
                        f"Bs. {_fmt_decimal_2(subtotal)}",
                    ]
                )

            details_table = Table(
                detail_rows,
                colWidths=[6.2 * cm, 2.2 * cm, 2.6 * cm, 2.7 * cm, 2.7 * cm],
                repeatRows=1,
            )
            details_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#374151")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D1D5DB")),
                        ("ALIGN", (3, 1), (4, -1), "RIGHT"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("TOPPADDING", (0, 0), (-1, -1), 3),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ]
                )
            )
            elements.append(details_table)
            elements.append(Spacer(1, 0.25 * cm))

    elements.append(Spacer(1, 0.35 * cm))
    elements.append(Paragraph("Sistema de Reportes", small_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def build_purchases_report_excel(
    *, purchases, from_date, to_date, include_items, generated_by, supplier_name=None
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Compras"

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)

    sheet["A1"] = "REPORTE DE COMPRAS"
    sheet["A1"].font = Font(size=14, bold=True)

    sheet["A3"] = "Rango"
    sheet["B3"] = f"{from_date} a {to_date}"
    sheet["A4"] = "Proveedor"
    sheet["B4"] = supplier_name or "Todos"
    sheet["A5"] = "Generado por"
    sheet["B5"] = generated_by
    sheet["A6"] = "Cantidad de compras"
    sheet["B6"] = purchases.count()
    total_amount = purchases.aggregate(total=Sum("total"))["total"] or Decimal("0.00")
    sheet["A7"] = "Total acumulado"
    sheet["B7"] = float(_fmt_decimal_2(total_amount))

    for cell_ref in ["A3", "A4", "A5", "A6", "A7"]:
        sheet[cell_ref].font = bold_font

    start_row = 9
    columns = ["Compra", "Fecha", "Proveedor", "Descripcion", "Total"]
    for idx, value in enumerate(columns, start=1):
        cell = sheet.cell(row=start_row, column=idx, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row = start_row + 1
    for purchase in purchases:
        sheet.cell(row=row, column=1, value=f"#{purchase.id}")
        sheet.cell(row=row, column=2, value=purchase.purchased_at.strftime("%Y-%m-%d %H:%M"))
        sheet.cell(
            row=row,
            column=3,
            value=purchase.supplier.business_name if purchase.supplier else "Sin proveedor",
        )
        sheet.cell(row=row, column=4, value=purchase.description or "-")
        sheet.cell(row=row, column=5, value=float(_fmt_decimal_2(purchase.total)))
        row += 1

        if include_items:
            sheet.cell(row=row, column=2, value=f"Detalle compra #{purchase.id}").font = bold_font
            row += 1

            detail_columns = ["Item", "Cantidad", "Unidad", "Unitario", "Subtotal"]
            for idx, value in enumerate(detail_columns, start=2):
                cell = sheet.cell(row=row, column=idx, value=value)
                cell.fill = PatternFill(
                    start_color="374151", end_color="374151", fill_type="solid"
                )
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            row += 1

            for item in purchase.items.all():
                subtotal = item.unit_price * item.quantity
                sheet.cell(row=row, column=2, value=item.raw_material.name)
                sheet.cell(row=row, column=3, value=float(_fmt_decimal_2(item.quantity)))
                sheet.cell(row=row, column=4, value=item.raw_material.measure_unit.name)
                sheet.cell(row=row, column=5, value=float(_fmt_decimal_2(item.unit_price)))
                sheet.cell(row=row, column=6, value=float(_fmt_decimal_2(subtotal)))
                row += 1

        row += 1

    sheet.column_dimensions["A"].width = 12
    sheet.column_dimensions["B"].width = 24
    sheet.column_dimensions["C"].width = 34
    sheet.column_dimensions["D"].width = 34
    sheet.column_dimensions["E"].width = 16
    sheet.column_dimensions["F"].width = 16

    for current_row in sheet.iter_rows(min_row=9, max_row=sheet.max_row, min_col=1, max_col=6):
        for cell in current_row:
            if cell.column in [5, 6] and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()
